"""
Rutas API para administración - Panel administrativo Phase 2
"""

import io
import logging
import os
from datetime import datetime, timezone
from functools import wraps

from flask import Blueprint, jsonify, request, send_file

from config import Config
from models.database import DatabaseManager, Segment, TranscriptionProject, User
from services.jwt_service import jwt_service
from services.reconstructed_transcript_service import ReconstructedTranscriptService

logger = logging.getLogger(__name__)

# Blueprint
admin_bp = Blueprint("admin_api", __name__, url_prefix="/api/v1/admin")

# Inicializar servicios
_db_manager = None
_reconstructed_service = None


def get_db_manager():
    global _db_manager
    if _db_manager is None:
        config = Config.from_env()
        _db_manager = DatabaseManager(config.DATABASE_URL)
    return _db_manager


def get_reconstructed_service():
    global _reconstructed_service
    if _reconstructed_service is None:
        config = Config.from_env()
        _reconstructed_service = ReconstructedTranscriptService(config=config)
    return _reconstructed_service


def serialize_user(user):
    return {"id": user.id, "username": user.username, "role": user.role}


# ==================== DECORADORES ====================


def admin_required(f):
    """Requiere ser admin"""

    @wraps(f)
    def decorated(*args, **kwargs):
        token = jwt_service.get_token_from_cookie_or_header()

        if not token:
            return jsonify({"error": "No token provided"}), 401

        try:
            payload = jwt_service.verify_access_token(token)
            if payload.get("role") != "admin":
                return jsonify({"error": "User is not admin"}), 403
            request.current_user = payload
            return f(*args, **kwargs)
        except Exception as e:
            logger.error(f"Admin token error: {e}")
            return jsonify({"error": "Invalid token"}), 401

    return decorated


# ==================== USUARIOS ====================


@admin_bp.route("/users", methods=["GET"])
@admin_required
def list_users():
    """Lista todos los usuarios"""
    try:
        db_manager = get_db_manager()
        session = db_manager.get_session()

        users = session.query(User).all()
        session.close()

        return (
            jsonify(
                {
                    "users": [serialize_user(u) for u in users]
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"Error listing users: {e}")
        return jsonify({"error": str(e)}), 500


@admin_bp.route("/users", methods=["POST"])
@admin_required
def create_user():
    """Crea un nuevo usuario"""
    try:
        data = request.get_json()
        username = data.get("username", "").strip()
        password = data.get("password", "")
        role = data.get("role", "annotator")

        if not username or len(username) < 3:
            return jsonify({"error": "Username must be at least 3 characters"}), 400

        if not password or len(password) < 6:
            return jsonify({"error": "Password must be at least 6 characters"}), 400

        if role not in ["admin", "annotator"]:
            return jsonify({"error": "Invalid role"}), 400

        db_manager = get_db_manager()
        session = db_manager.get_session()

        # Verificar si existe
        existing = session.query(User).filter_by(username=username).first()
        if existing:
            session.close()
            return jsonify({"error": "User already exists"}), 409

        # Crear usuario
        user = User(username=username, password=password, role=role)
        session.add(user)
        session.commit()

        user_data = serialize_user(user)
        session.close()

        return jsonify({"message": f"User {username} created", "user": user_data}), 201

    except Exception as e:
        logger.error(f"Error creating user: {e}")
        return jsonify({"error": str(e)}), 500


@admin_bp.route("/users/<int:user_id>", methods=["PUT"])
@admin_required
def update_user(user_id):
    """Actualiza username, password y/o rol de un usuario."""
    session = None
    try:
        data = request.get_json() or {}
        username = (data.get("username") or "").strip()
        password = data.get("password")
        role = data.get("role")

        if not username and password is None and role is None:
            return jsonify({"error": "No changes provided"}), 400

        if username and len(username) < 3:
            return jsonify({"error": "Username must be at least 3 characters"}), 400

        if password is not None and len(password) < 6:
            return jsonify({"error": "Password must be at least 6 characters"}), 400

        if role is not None and role not in ["admin", "annotator"]:
            return jsonify({"error": "Invalid role"}), 400

        db_manager = get_db_manager()
        session = db_manager.get_session()

        user = session.query(User).filter_by(id=user_id).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        if username and username != user.username:
            existing = session.query(User).filter(User.username == username, User.id != user_id).first()
            if existing:
                return jsonify({"error": "User already exists"}), 409
            user.username = username

        if password:
            user.set_password(password)

        if role and role != user.role:
            if user_id == request.current_user["user_id"] and role != "admin":
                return jsonify({"error": "Cannot remove your own admin role"}), 403

            if user.role == "admin" and role != "admin":
                admin_count = session.query(User).filter_by(role="admin").count()
                if admin_count <= 1:
                    return jsonify({"error": "Cannot demote the last admin"}), 403

            user.role = role

        session.commit()
        user_data = serialize_user(user)
        return jsonify({"message": f"User {user.username} updated", "user": user_data}), 200

    except Exception as e:
        if session is not None:
            session.rollback()
        logger.error(f"Error updating user: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if session is not None:
            session.close()


@admin_bp.route("/users/<int:user_id>", methods=["DELETE"])
@admin_required
def delete_user(user_id):
    """Elimina un usuario"""
    session = None
    try:
        if user_id == request.current_user["user_id"]:
            return jsonify({"error": "Cannot delete yourself"}), 403

        db_manager = get_db_manager()
        session = db_manager.get_session()

        user = session.query(User).filter_by(id=user_id).first()
        if not user:
            return jsonify({"error": "User not found"}), 404

        if user.role == "admin":
            admin_count = session.query(User).filter_by(role="admin").count()
            if admin_count <= 1:
                return jsonify({"error": "Cannot delete the last admin"}), 403

        affected_segments = session.query(Segment).filter_by(annotator_id=user_id).all()
        for segment in affected_segments:
            segment.annotator_id = None
            if segment.review_status == "pending":
                segment.text_revised = None
                segment.completed_at = None
            segment.updated_at = datetime.now(timezone.utc)

        username = user.username
        session.delete(user)
        session.commit()

        return (
            jsonify(
                {
                    "message": f"User {username} deleted",
                    "released_segments": len(affected_segments),
                }
            ),
            200,
        )

    except Exception as e:
        if session is not None:
            session.rollback()
        logger.error(f"Error deleting user: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if session is not None:
            session.close()


@admin_bp.route("/users/<int:user_id>/stats", methods=["GET"])
@admin_required
def get_user_stats(user_id):
    """Obtiene estadísticas de un usuario"""
    try:
        db_manager = get_db_manager()
        session = db_manager.get_session()

        user = session.query(User).filter_by(id=user_id).first()
        if not user:
            session.close()
            return jsonify({"error": "User not found"}), 404

        # Contar segmentos asignados y completados
        total_assigned = session.query(Segment).filter_by(annotator_id=user_id).count()
        completed = (
            session.query(Segment)
            .filter_by(annotator_id=user_id)
            .filter(Segment.review_status.in_(["approved", "corrected"]))
            .count()
        )

        session.close()

        return (
            jsonify(
                {
                    "user_id": user_id,
                    "username": user.username,
                    "total_segments": total_assigned,
                    "approved_segments": completed,
                    "corrected_segments": completed,
                    "pending_segments": total_assigned - completed,
                    "words_completed_percentage": (
                        round((completed / total_assigned * 100), 2)
                        if total_assigned > 0
                        else 0
                    ),
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"Error getting user stats: {e}")
        return jsonify({"error": str(e)}), 500


# ==================== PROYECTOS ====================


@admin_bp.route("/projects", methods=["GET"])
@admin_required
def list_projects():
    """Lista todos los proyectos"""
    try:
        db_manager = get_db_manager()
        session = db_manager.get_session()

        projects = session.query(TranscriptionProject).all()
        session.close()

        return (
            jsonify(
                {
                    "projects": [
                        {"id": p.id, "name": p.name, "description": p.description}
                        for p in projects
                    ]
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"Error listing projects: {e}")
        return jsonify({"error": str(e)}), 500


@admin_bp.route("/projects/<project_id>/stats", methods=["GET"])
@admin_required
def get_project_stats(project_id):
    """Obtiene estadísticas de un proyecto"""
    try:
        db_manager = get_db_manager()
        session = db_manager.get_session()

        project = session.query(TranscriptionProject).filter_by(id=project_id).first()
        if not project:
            session.close()
            return jsonify({"error": "Project not found"}), 404

        total_segments = session.query(Segment).filter_by(project_id=project_id).count()
        pending = (
            session.query(Segment)
            .filter_by(project_id=project_id, review_status="pending")
            .count()
        )
        approved = (
            session.query(Segment)
            .filter_by(project_id=project_id, review_status="approved")
            .count()
        )
        corrected = (
            session.query(Segment)
            .filter_by(project_id=project_id, review_status="corrected")
            .count()
        )

        # Contar anotadores
        from sqlalchemy import distinct, func

        total_annotators = (
            session.query(func.count(distinct(Segment.annotator_id)))
            .filter_by(project_id=project_id)
            .scalar()
            or 0
        )

        session.close()

        return (
            jsonify(
                {
                    "project_id": project_id,
                    "total_segments": total_segments,
                    "pending_segments": pending,
                    "approved_segments": approved,
                    "corrected_segments": corrected,
                    "words_completed_percentage": (
                        round(((approved + corrected) / total_segments * 100), 2)
                        if total_segments > 0
                        else 0
                    ),
                    "total_annotators": total_annotators,
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"Error getting project stats: {e}")
        return jsonify({"error": str(e)}), 500


# ==================== SEGMENTOS ====================


@admin_bp.route("/projects/<project_id>/segments", methods=["GET"])
@admin_required
def list_segments(project_id):
    """Lista segmentos de un proyecto"""
    try:
        status = request.args.get("status", "all")

        db_manager = get_db_manager()
        session = db_manager.get_session()

        query = session.query(Segment).filter_by(project_id=project_id)

        if status != "all":
            if status == "pending":
                query = query.filter_by(review_status="pending").filter(
                    Segment.annotator_id.is_(None)
                )
            elif status == "completed":
                query = query.filter(
                    Segment.review_status.in_(["approved", "corrected"])
                )

        segments = query.all()
        session.close()

        return (
            jsonify(
                {
                    "words": [  # Mantener nombre 'words' por compatibility
                        {
                            "id": s.id,
                            "text": s.text,
                            "status": s.review_status,
                            "assigned_to": s.annotator_id,
                            "created_at": (
                                s.created_at.isoformat() if s.created_at else None
                            ),
                        }
                        for s in segments
                    ]
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"Error listing segments: {e}")
        return jsonify({"error": str(e)}), 500


@admin_bp.route(
    "/projects/<project_id>/segments/<int:segment_id>/assign", methods=["POST"]
)
@admin_required
def assign_segment(project_id, segment_id):
    """Asigna un segmento a un anotador"""
    try:
        data = request.get_json()
        annotator_id = data.get("annotator_id")

        if not annotator_id:
            return jsonify({"error": "annotator_id required"}), 400

        db_manager = get_db_manager()
        session = db_manager.get_session()

        segment = (
            session.query(Segment)
            .filter_by(id=segment_id, project_id=project_id)
            .first()
        )

        if not segment:
            session.close()
            return jsonify({"error": "Segment not found"}), 404

        segment.annotator_id = annotator_id
        segment.updated_at = datetime.now(timezone.utc)
        session.commit()
        session.close()

        return jsonify({"message": "Segment assigned", "segment_id": segment_id}), 200

    except Exception as e:
        logger.error(f"Error assigning segment: {e}")
        return jsonify({"error": str(e)}), 500


@admin_bp.route("/projects/<project_id>/assign-by-count", methods=["POST"])
@admin_required
def assign_segments_by_count(project_id):
    """Asigna una cantidad de segmentos pendientes a un anotador."""
    session = None
    try:
        data = request.get_json() or {}
        annotator_id = data.get("annotator_id")
        segment_count = data.get("segment_count")

        if not annotator_id:
            return jsonify({"error": "annotator_id required"}), 400

        try:
            segment_count = int(segment_count)
        except (TypeError, ValueError):
            return jsonify({"error": "segment_count must be an integer"}), 400

        if segment_count <= 0:
            return jsonify({"error": "segment_count must be greater than 0"}), 400

        db_manager = get_db_manager()
        session = db_manager.get_session()

        annotator = session.query(User).filter_by(id=annotator_id).first()
        if not annotator:
            return jsonify({"error": "Annotator not found"}), 404

        segments = (
            session.query(Segment)
            .filter_by(project_id=project_id, review_status="pending", annotator_id=None)
            .order_by(Segment.audio_filename.asc(), Segment.segment_index.asc())
            .limit(segment_count)
            .all()
        )

        if not segments:
            return jsonify({"error": "No pending segments available"}), 404

        assigned_ids = []
        for segment in segments:
            segment.annotator_id = annotator_id
            segment.updated_at = datetime.now(timezone.utc)
            assigned_ids.append(segment.id)

        session.commit()

        return (
            jsonify(
                {
                    "message": f"Assigned {len(assigned_ids)} segment(s)",
                    "annotator_id": annotator_id,
                    "segment_ids": assigned_ids,
                }
            ),
            200,
        )

    except Exception as e:
        if session is not None:
            session.rollback()
        logger.error(f"Error assigning segments by count: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if session is not None:
            session.close()


@admin_bp.route(
    "/projects/<project_id>/segments/<int:segment_id>/unassign", methods=["POST"]
)
@admin_required
def unassign_segment(project_id, segment_id):
    """Desasigna un segmento (lo devuelve al pool de asignables)"""
    try:
        db_manager = get_db_manager()
        session = db_manager.get_session()

        segment = (
            session.query(Segment)
            .filter_by(id=segment_id, project_id=project_id)
            .first()
        )

        if not segment:
            session.close()
            return jsonify({"error": "Segment not found"}), 404

        segment.annotator_id = None
        segment.review_status = "pending"
        segment.text_revised = None
        segment.completed_at = None
        segment.updated_at = datetime.now(timezone.utc)
        session.commit()
        session.close()

        return jsonify({"message": "Segment unassigned", "segment_id": segment_id}), 200

    except Exception as e:
        logger.error(f"Error unassigning segment: {e}")
        return jsonify({"error": str(e)}), 500


@admin_bp.route("/projects/<project_id>/assigned", methods=["GET"])
@admin_required
def list_assigned_segments(project_id):
    """Lista segmentos asignados agrupados por usuario"""
    try:
        db_manager = get_db_manager()
        session = db_manager.get_session()

        # Obtener todos los segmentos asignados
        segments = (
            session.query(Segment)
            .filter(Segment.project_id == project_id, Segment.annotator_id.isnot(None))
            .all()
        )

        # Agrupar por usuario
        by_user = {}
        for s in segments:
            uid = s.annotator_id
            if uid not in by_user:
                user = session.query(User).filter_by(id=uid).first()
                by_user[uid] = {
                    "user_id": uid,
                    "username": user.username if user else f"user_{uid}",
                    "segments": [],
                }
            by_user[uid]["segments"].append(
                {
                    "id": s.id,
                    "text": s.text,
                    "status": s.review_status,
                    "completed_at": (
                        s.completed_at.isoformat() if s.completed_at else None
                    ),
                }
            )

        session.close()

        return jsonify({"assigned": list(by_user.values())}), 200

    except Exception as e:
        logger.error(f"Error listing assigned segments: {e}")
        return jsonify({"error": str(e)}), 500


# ==================== ESTADÍSTICAS GLOBALES ====================


@admin_bp.route("/users/<int:user_id>/annotations", methods=["GET"])
@admin_required
def get_user_annotations(user_id):
    """Lista todas las anotaciones (segmentos completados) de un usuario"""
    try:
        db_manager = get_db_manager()
        session = db_manager.get_session()

        user = session.query(User).filter_by(id=user_id).first()
        if not user:
            session.close()
            return jsonify({"error": "User not found"}), 404

        segments = (
            session.query(Segment)
            .filter(
                Segment.annotator_id == user_id,
                Segment.review_status.in_(["approved", "corrected"]),
            )
            .order_by(Segment.completed_at.desc())
            .all()
        )

        result = []
        for s in segments:
            result.append(
                {
                    "id": s.id,
                    "project_id": s.project_id,
                    "segment_index": s.segment_index,
                    "text": s.text,
                    "text_revised": s.text_revised,
                    "review_status": s.review_status,
                    "start_time": s.start_time,
                    "end_time": s.end_time,
                    "completed_at": (
                        s.completed_at.isoformat() if s.completed_at else None
                    ),
                }
            )

        session.close()

        return (
            jsonify(
                {
                    "user_id": user_id,
                    "username": user.username,
                    "annotations": result,
                    "total": len(result),
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"Error getting user annotations: {e}")
        return jsonify({"error": str(e)}), 500


@admin_bp.route("/annotations/export", methods=["GET"])
@admin_required
def export_all_annotations():
    """Exporta todas las anotaciones completadas de todos los usuarios a Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        db_manager = get_db_manager()
        session = db_manager.get_session()

        segments = (
            session.query(Segment)
            .filter(Segment.review_status.in_(["approved", "corrected"]))
            .order_by(Segment.completed_at.desc())
            .all()
        )

        # Pre-load annotator usernames
        annotator_ids = set(s.annotator_id for s in segments if s.annotator_id)
        users_map = {}
        if annotator_ids:
            users = session.query(User).filter(User.id.in_(annotator_ids)).all()
            users_map = {u.id: u.username for u in users}

        data_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "data", "transcription_projects"
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Anotaciones"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = [
            "Ruta del Audio",
            "Timestamp (inicio - fin)",
            "Segmento Original",
            "Segmento Modificado",
            "Anotador",
            "Fecha de Anotación",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        def fmt_time(seconds):
            m, sec = divmod(seconds, 60)
            h, m = divmod(int(m), 60)
            return f"{h:02d}:{int(m):02d}:{sec:05.2f}"

        for row_idx, s in enumerate(segments, 2):
            audio_path = os.path.join(data_dir, s.project_id, s.audio_filename)
            timestamp = f"{fmt_time(s.start_time)} - {fmt_time(s.end_time)}"
            completed = (
                s.completed_at.strftime("%Y-%m-%d %H:%M")
                if s.completed_at
                else "\u2014"
            )
            username = users_map.get(s.annotator_id, "\u2014")

            row_data = [
                audio_path,
                timestamp,
                s.text,
                s.text_revised or s.text,
                username,
                completed,
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 45
        ws.column_dimensions["D"].width = 45
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 20

        session.close()

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'anotaciones_todas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        logger.error(f"Error exporting annotations: {e}")
        return jsonify({"error": str(e)}), 500


@admin_bp.route("/annotations/<int:segment_id>", methods=["PUT"])
@admin_required
def edit_annotation(segment_id):
    """Permite al admin editar el texto revisado de una anotación"""
    try:
        data = request.get_json()
        text_revised = data.get("text_revised")
        review_status = data.get("review_status")
        decision_type = data.get("decision_type")

        if not text_revised:
            return jsonify({"error": "text_revised required"}), 400

        db_manager = get_db_manager()
        session = db_manager.get_session()

        segment = session.query(Segment).filter_by(id=segment_id).first()
        if not segment:
            session.close()
            return jsonify({"error": "Segment not found"}), 404

        segment.text_revised = text_revised
        if review_status in ("approved", "corrected"):
            segment.review_status = review_status
        if decision_type in ("confirmed", "doubtful", "discarded"):
            segment.decision_type = decision_type
        segment.updated_at = datetime.now(timezone.utc)
        session.flush()
        export_result = get_reconstructed_service().write_project_exports(
            session, segment.project_id
        )
        session.commit()
        session.close()

        return (
            jsonify(
                {
                    "message": "Annotation updated",
                    "segment_id": segment_id,
                    "exports": export_result,
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"Error editing annotation: {e}")
        return jsonify({"error": str(e)}), 500


@admin_bp.route("/annotations/<int:segment_id>/revert", methods=["POST"])
@admin_required
def revert_annotation(segment_id):
    """Revierte una anotación: la vuelve a estado pending sin anotador"""
    try:
        db_manager = get_db_manager()
        session = db_manager.get_session()

        segment = session.query(Segment).filter_by(id=segment_id).first()
        if not segment:
            session.close()
            return jsonify({"error": "Segment not found"}), 404

        segment.annotator_id = None
        segment.review_status = "pending"
        segment.decision_type = None
        segment.text_revised = None
        segment.completed_at = None
        segment.updated_at = datetime.now(timezone.utc)
        session.flush()
        export_result = get_reconstructed_service().write_project_exports(
            session, segment.project_id
        )
        session.commit()
        session.close()

        return (
            jsonify(
                {
                    "message": "Annotation reverted",
                    "segment_id": segment_id,
                    "exports": export_result,
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"Error reverting annotation: {e}")
        return jsonify({"error": str(e)}), 500


@admin_bp.route("/annotations/bulk-revert", methods=["POST"])
@admin_required
def bulk_revert_annotations():
    """Revierte múltiples anotaciones a estado pending"""
    try:
        data = request.get_json()
        segment_ids = data.get("segment_ids", [])

        if not segment_ids:
            return jsonify({"error": "segment_ids required"}), 400

        db_manager = get_db_manager()
        session = db_manager.get_session()

        reverted = 0
        affected_projects = set()
        for sid in segment_ids:
            segment = session.query(Segment).filter_by(id=sid).first()
            if segment:
                affected_projects.add(segment.project_id)
                segment.annotator_id = None
                segment.review_status = "pending"
                segment.decision_type = None
                segment.text_revised = None
                segment.completed_at = None
                segment.updated_at = datetime.now(timezone.utc)
                reverted += 1

        session.flush()
        export_results = {
            project_id: get_reconstructed_service().write_project_exports(
                session, project_id
            )
            for project_id in affected_projects
        }
        session.commit()
        session.close()

        return (
            jsonify(
                {
                    "message": f"{reverted} annotations reverted",
                    "reverted": reverted,
                    "exports": export_results,
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"Error bulk reverting annotations: {e}")
        return jsonify({"error": str(e)}), 500


@admin_bp.route("/stats", methods=["GET"])
@admin_required
def get_global_stats():
    """Obtiene estadísticas globales del sistema"""
    try:
        db_manager = get_db_manager()
        session = db_manager.get_session()

        total_users = session.query(User).count()
        admin_count = session.query(User).filter_by(role="admin").count()
        annotator_count = total_users - admin_count

        total_projects = session.query(TranscriptionProject).count()
        total_segments = session.query(Segment).count()
        completed_segments = (
            session.query(Segment)
            .filter(Segment.review_status.in_(["approved", "corrected"]))
            .count()
        )

        session.close()

        return (
            jsonify(
                {
                    "total_users": total_users,
                    "admins": admin_count,
                    "annotators": annotator_count,
                    "total_projects": total_projects,
                    "total_segments": total_segments,
                    "completed_segments": completed_segments,
                    "completion_percentage": (
                        round((completed_segments / total_segments * 100), 2)
                        if total_segments > 0
                        else 0
                    ),
                }
            ),
            200,
        )

    except Exception as e:
        logger.error(f"Error getting global stats: {e}")
        return jsonify({"error": str(e)}), 500
